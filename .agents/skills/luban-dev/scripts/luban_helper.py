#!/usr/bin/env python
"""Tyou Luban Excel helper.

Default mode is read-only. Commands that modify xlsx files require --write.
This helper targets Tyou's Design/config layout and #*.xlsx auto-import tables.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import Workbook, load_workbook


ROOT = Path.cwd()
DEFAULT_DATA_DIR = ROOT / "Design" / "config"
GENERATED_SCHEMA = ROOT / "Client" / "assets" / "scripts" / "proto" / "config" / "bin" / "schema.ts"


def die(message: str) -> None:
    print(f"ERROR: {message}", file=sys.stderr)
    raise SystemExit(1)


def require_write(args: argparse.Namespace) -> None:
    if not getattr(args, "write", False):
        die("write operation refused; pass --write after confirming the OpenSpec task and references")


def data_dir(args: argparse.Namespace) -> Path:
    return Path(args.data_dir).resolve()


def load_json_arg(value: str) -> Any:
    path = Path(value)
    if path.exists():
        return json.loads(path.read_text(encoding="utf-8"))
    try:
        return json.loads(value)
    except json.JSONDecodeError:
        result: Dict[str, Any] = {}
        for part in value.split(","):
            if "=" not in part:
                raise
            key, raw = part.split("=", 1)
            raw = raw.strip()
            if raw.isdigit():
                parsed: Any = int(raw)
            else:
                try:
                    parsed = float(raw)
                except ValueError:
                    parsed = raw
            result[key.strip()] = parsed
        return result


def parse_fields(value: str) -> List[Dict[str, Any]]:
    if not value:
        return []
    parsed: List[Dict[str, Any]] = []
    for part in value.split(","):
        chunks = part.split(":", 2)
        parsed.append(
            {
                "name": chunks[0].strip(),
                "type": chunks[1].strip() if len(chunks) > 1 else "string",
                "comment": chunks[2].strip() if len(chunks) > 2 else "",
            }
        )
    return parsed


def parse_enum_items(value: str) -> List[Dict[str, Any]]:
    if not value:
        return []
    parsed: List[Dict[str, Any]] = []
    for part in value.split(","):
        name_part, _, comment = part.partition(":")
        name, _, raw_value = name_part.partition("=")
        item: Dict[str, Any] = {"name": name.strip(), "alias": comment.strip()}
        if raw_value.strip():
            item["value"] = int(raw_value.strip())
        parsed.append(item)
    return parsed


def cell_value(value: Any) -> Any:
    return "" if value is None else value


class AutoTable:
    def __init__(self, path: Path):
        self.path = path
        self.name = path.stem[1:] if path.stem.startswith("#") else path.stem

    @property
    def file_name(self) -> str:
        return self.path.name

    def workbook(self):
        return load_workbook(self.path)

    def sheet(self):
        wb = self.workbook()
        return wb, wb[wb.sheetnames[0]]


class LubanHelper:
    def __init__(self, data_dir_path: Path):
        self.data_dir = data_dir_path
        if not self.data_dir.exists():
            die(f"data dir not found: {self.data_dir}")

    def auto_tables(self) -> List[AutoTable]:
        return [AutoTable(p) for p in sorted(self.data_dir.glob("#*.xlsx")) if p.is_file()]

    def find_table(self, name: str) -> Optional[AutoTable]:
        lowered = name.lower()
        for table in self.auto_tables():
            if table.name.lower() == lowered or table.file_name.lower() == lowered or f"#{table.name}".lower() == lowered:
                return table
        return None

    def list_tables(self) -> List[Dict[str, Any]]:
        rows: List[Dict[str, Any]] = []
        for table in self.auto_tables():
            wb, ws = table.sheet()
            rows.append(
                {
                    "name": table.name,
                    "file": table.file_name,
                    "sheet": ws.title,
                    "rows": max(ws.max_row - 3, 0),
                    "columns": ws.max_column,
                }
            )
            wb.close()
        return rows

    def table_info(self, name: str) -> Dict[str, Any]:
        table = self.find_table(name)
        if not table:
            die(f"table not found: {name}")
        wb, ws = table.sheet()
        fields = self.fields(table.name)
        data_rows = self.rows(table.name, limit=5)
        info = {
            "name": table.name,
            "file": table.file_name,
            "sheet": ws.title,
            "rows": max(ws.max_row - 3, 0),
            "columns": ws.max_column,
            "fields": fields,
            "sample": data_rows,
        }
        wb.close()
        return info

    def add_table(self, name: str, fields: List[Dict[str, Any]]) -> None:
        clean = name[1:] if name.startswith("#") else name
        path = self.data_dir / f"#{clean}.xlsx"
        if path.exists():
            die(f"table already exists: {path.name}")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.cell(1, 1).value = "##var"
        ws.cell(2, 1).value = "##type"
        ws.cell(3, 1).value = "##"
        for idx, field in enumerate(fields, start=2):
            ws.cell(1, idx).value = field["name"]
            ws.cell(2, idx).value = field.get("type", "string")
            ws.cell(3, idx).value = field.get("comment", "")
        wb.save(path)
        wb.close()

    def delete_table(self, name: str) -> None:
        table = self.find_table(name)
        if not table:
            die(f"table not found: {name}")
        table.path.unlink()

    def fields(self, name: str) -> List[Dict[str, Any]]:
        table = self.find_table(name)
        if not table:
            die(f"table not found: {name}")
        wb, ws = table.sheet()
        fields: List[Dict[str, Any]] = []
        for col in range(2, ws.max_column + 1):
            field_name = cell_value(ws.cell(1, col).value)
            if field_name == "":
                continue
            fields.append(
                {
                    "name": field_name,
                    "type": cell_value(ws.cell(2, col).value),
                    "comment": cell_value(ws.cell(3, col).value),
                    "column": col,
                }
            )
        wb.close()
        return fields

    def rows(self, name: str, start: int = 0, limit: int = 20) -> List[Dict[str, Any]]:
        table = self.find_table(name)
        if not table:
            die(f"table not found: {name}")
        wb, ws = table.sheet()
        headers = [cell_value(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
        output: List[Dict[str, Any]] = []
        first = 4 + start
        last = min(ws.max_row, first + limit - 1)
        for row_idx in range(first, last + 1):
            row: Dict[str, Any] = {"_row": row_idx}
            for col, header in enumerate(headers, start=1):
                if header:
                    row[header] = cell_value(ws.cell(row_idx, col).value)
            output.append(row)
        wb.close()
        return output

    def query_rows(self, name: str, conditions: Dict[str, Any], limit: int = 20) -> List[Dict[str, Any]]:
        table = self.find_table(name)
        if not table:
            die(f"table not found: {name}")
        wb, ws = table.sheet()
        headers = [cell_value(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
        index = {h: i + 1 for i, h in enumerate(headers) if h}
        missing = [key for key in conditions if key not in index]
        if missing:
            die(f"unknown fields: {', '.join(missing)}")
        output: List[Dict[str, Any]] = []
        for row_idx in range(4, ws.max_row + 1):
            ok = True
            for key, expected in conditions.items():
                if str(cell_value(ws.cell(row_idx, index[key]).value)) != str(expected):
                    ok = False
                    break
            if ok:
                row = {"_row": row_idx}
                for col, header in enumerate(headers, start=1):
                    if header:
                        row[header] = cell_value(ws.cell(row_idx, col).value)
                output.append(row)
                if len(output) >= limit:
                    break
        wb.close()
        return output

    def add_field(self, name: str, field_name: str, field_type: str, comment: str) -> None:
        table = self.find_table(name)
        if not table:
            die(f"table not found: {name}")
        wb, ws = table.sheet()
        existing = [str(cell_value(ws.cell(1, col).value)) for col in range(1, ws.max_column + 1)]
        if field_name in existing:
            die(f"field already exists: {field_name}")
        col = ws.max_column + 1
        ws.cell(1, col).value = field_name
        ws.cell(2, col).value = field_type
        ws.cell(3, col).value = comment
        wb.save(table.path)
        wb.close()

    def update_field(self, name: str, field_name: str, new_name: str = "", field_type: str = "", comment: str = "") -> None:
        table = self.find_table(name)
        if not table:
            die(f"table not found: {name}")
        wb, ws = table.sheet()
        col = self._field_col(ws, field_name)
        if not col:
            die(f"field not found: {field_name}")
        if new_name:
            ws.cell(1, col).value = new_name
        if field_type:
            ws.cell(2, col).value = field_type
        if comment:
            ws.cell(3, col).value = comment
        wb.save(table.path)
        wb.close()

    def delete_field(self, name: str, field_name: str) -> None:
        table = self.find_table(name)
        if not table:
            die(f"table not found: {name}")
        wb, ws = table.sheet()
        col = self._field_col(ws, field_name)
        if not col:
            die(f"field not found: {field_name}")
        ws.delete_cols(col, 1)
        wb.save(table.path)
        wb.close()

    def add_row(self, name: str, data: Dict[str, Any]) -> None:
        table = self.find_table(name)
        if not table:
            die(f"table not found: {name}")
        wb, ws = table.sheet()
        headers = [cell_value(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
        index = {h: i + 1 for i, h in enumerate(headers) if h}
        unknown = [key for key in data if key not in index]
        if unknown:
            die(f"unknown fields: {', '.join(unknown)}")
        row_idx = ws.max_row + 1
        for key, value in data.items():
            ws.cell(row_idx, index[key]).value = value
        wb.save(table.path)
        wb.close()

    def update_row(self, name: str, row_index: int, data: Dict[str, Any]) -> None:
        table = self.find_table(name)
        if not table:
            die(f"table not found: {name}")
        wb, ws = table.sheet()
        if row_index < 4 or row_index > ws.max_row:
            die("row index must be an Excel row number within data rows")
        headers = [cell_value(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
        index = {h: i + 1 for i, h in enumerate(headers) if h}
        unknown = [key for key in data if key not in index]
        if unknown:
            die(f"unknown fields: {', '.join(unknown)}")
        for key, value in data.items():
            ws.cell(row_index, index[key]).value = value
        wb.save(table.path)
        wb.close()

    def delete_row(self, name: str, row_index: int) -> None:
        table = self.find_table(name)
        if not table:
            die(f"table not found: {name}")
        wb, ws = table.sheet()
        if row_index < 4 or row_index > ws.max_row:
            die("row index must be an Excel row number within data rows")
        ws.delete_rows(row_index, 1)
        wb.save(table.path)
        wb.close()

    def validate_table(self, name: str) -> Dict[str, Any]:
        table = self.find_table(name)
        if not table:
            die(f"table not found: {name}")
        wb, ws = table.sheet()
        errors: List[str] = []
        fields = self.fields(table.name)
        if not fields:
            errors.append("no fields")
        names = [f["name"] for f in fields]
        duplicates = sorted({n for n in names if names.count(n) > 1})
        for dup in duplicates:
            errors.append(f"duplicate field: {dup}")
        id_col = self._field_col(ws, "id")
        if id_col:
            seen = {}
            for row_idx in range(4, ws.max_row + 1):
                value = cell_value(ws.cell(row_idx, id_col).value)
                if value == "":
                    errors.append(f"empty id at row {row_idx}")
                elif value in seen:
                    errors.append(f"duplicate id {value} at rows {seen[value]} and {row_idx}")
                seen[value] = row_idx
        wb.close()
        return {"table": table.name, "ok": len(errors) == 0, "errors": errors}

    def validate_all(self) -> Dict[str, Any]:
        results = [self.validate_table(t.name) for t in self.auto_tables()]
        return {"ok": all(r["ok"] for r in results), "results": results}

    def ref(self, name: str) -> Dict[str, Any]:
        pattern = re.escape(name)
        roots = [ROOT / "Client" / "assets", self.data_dir]
        matches: List[Dict[str, Any]] = []
        for root in roots:
            if not root.exists():
                continue
            for path in root.rglob("*"):
                if not path.is_file() or path.suffix.lower() in {".png", ".jpg", ".bin"}:
                    continue
                try:
                    text = path.read_text(encoding="utf-8", errors="ignore")
                except Exception:
                    continue
                for line_no, line in enumerate(text.splitlines(), start=1):
                    if re.search(pattern, line):
                        matches.append(
                            {
                                "path": str(path.relative_to(ROOT)).replace("\\", "/"),
                                "line": line_no,
                                "text": line.strip()[:200],
                            }
                        )
        return {"query": name, "matches": matches[:200], "count": len(matches)}

    def list_enums(self) -> List[Dict[str, Any]]:
        return self._list_definition_rows("__enums__.xlsx", key="full_name")

    def get_enum(self, name: str) -> Optional[Dict[str, Any]]:
        return self._get_definition_group("__enums__.xlsx", name)

    def add_enum(self, name: str, items: List[Dict[str, Any]], comment: str = "", flags: bool = False) -> None:
        path = self.data_dir / "__enums__.xlsx"
        wb = load_workbook(path)
        ws = wb[wb.sheetnames[0]]
        if self.get_enum(name):
            wb.close()
            die(f"enum already exists: {name}")
        start = ws.max_row + 1
        for offset, item in enumerate(items):
            row = start + offset
            ws.cell(row, 2).value = name if offset == 0 else None
            ws.cell(row, 3).value = "1" if flags and offset == 0 else None
            ws.cell(row, 6).value = comment if offset == 0 else None
            ws.cell(row, 8).value = item["name"]
            ws.cell(row, 9).value = item.get("alias", "")
            ws.cell(row, 10).value = item.get("value", offset)
        wb.save(path)
        wb.close()

    def update_enum(self, name: str, comment: str = "", flags: Optional[bool] = None) -> None:
        path = self.data_dir / "__enums__.xlsx"
        wb = load_workbook(path)
        ws = wb[wb.sheetnames[0]]
        row = self._definition_start_row(ws, name)
        if not row:
            wb.close()
            die(f"enum not found: {name}")
        if flags is not None:
            ws.cell(row, 3).value = "1" if flags else None
        if comment:
            ws.cell(row, 6).value = comment
        wb.save(path)
        wb.close()

    def delete_enum(self, name: str) -> None:
        self._delete_definition_group("__enums__.xlsx", name)

    def list_beans(self) -> List[Dict[str, Any]]:
        return self._list_definition_rows("__beans__.xlsx", key="full_name")

    def get_bean(self, name: str) -> Optional[Dict[str, Any]]:
        return self._get_definition_group("__beans__.xlsx", name)

    def add_bean(self, name: str, fields: List[Dict[str, Any]], parent: str = "", comment: str = "") -> None:
        path = self.data_dir / "__beans__.xlsx"
        wb = load_workbook(path)
        ws = wb[wb.sheetnames[0]]
        if self.get_bean(name):
            wb.close()
            die(f"bean already exists: {name}")
        start = ws.max_row + 1
        if not fields:
            fields = [{"name": "", "type": "", "comment": ""}]
        for offset, field in enumerate(fields):
            row = start + offset
            ws.cell(row, 2).value = name if offset == 0 else None
            ws.cell(row, 3).value = parent if offset == 0 else None
            ws.cell(row, 7).value = comment if offset == 0 else None
            ws.cell(row, 10).value = field.get("name", "")
            ws.cell(row, 11).value = field.get("type", "string")
            ws.cell(row, 12).value = field.get("comment", "")
        wb.save(path)
        wb.close()

    def update_bean(self, name: str, parent: str = "", comment: str = "") -> None:
        path = self.data_dir / "__beans__.xlsx"
        wb = load_workbook(path)
        ws = wb[wb.sheetnames[0]]
        row = self._definition_start_row(ws, name)
        if not row:
            wb.close()
            die(f"bean not found: {name}")
        if parent:
            ws.cell(row, 3).value = parent
        if comment:
            ws.cell(row, 7).value = comment
        wb.save(path)
        wb.close()

    def delete_bean(self, name: str) -> None:
        self._delete_definition_group("__beans__.xlsx", name)

    def _list_definition_rows(self, file_name: str, key: str) -> List[Dict[str, Any]]:
        path = self.data_dir / file_name
        if not path.exists():
            return []
        wb = load_workbook(path, data_only=False)
        rows: List[Dict[str, Any]] = []
        for ws in wb.worksheets:
            headers = [cell_value(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
            for row_idx in range(4, ws.max_row + 1):
                row = {"_sheet": ws.title, "_row": row_idx}
                for col, header in enumerate(headers, start=1):
                    if header:
                        row[header] = cell_value(ws.cell(row_idx, col).value)
                if row.get(key):
                    rows.append(row)
        wb.close()
        return rows

    def _get_definition(self, file_name: str, name: str) -> Optional[Dict[str, Any]]:
        lowered = name.lower()
        for row in self._list_definition_rows(file_name, "full_name"):
            full_name = str(row.get("full_name", ""))
            if full_name.lower() == lowered or full_name.split(".")[-1].lower() == lowered:
                return row
        return None

    def _get_definition_group(self, file_name: str, name: str) -> Optional[Dict[str, Any]]:
        path = self.data_dir / file_name
        if not path.exists():
            return None
        wb = load_workbook(path, data_only=False)
        lowered = name.lower()
        for ws in wb.worksheets:
            headers = [cell_value(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
            row = 4
            while row <= ws.max_row:
                full_name = str(cell_value(ws.cell(row, 2).value))
                if full_name and (full_name.lower() == lowered or full_name.split(".")[-1].lower() == lowered):
                    rows: List[Dict[str, Any]] = []
                    scan = row
                    while scan <= ws.max_row:
                        if scan != row and cell_value(ws.cell(scan, 2).value):
                            break
                        item = {"_sheet": ws.title, "_row": scan}
                        for col, header in enumerate(headers, start=1):
                            if header:
                                item[header] = cell_value(ws.cell(scan, col).value)
                        rows.append(item)
                        scan += 1
                    wb.close()
                    return {"full_name": full_name, "rows": rows}
                row += 1
        wb.close()
        return None

    def _definition_start_row(self, ws, name: str) -> int:
        lowered = name.lower()
        for row in range(4, ws.max_row + 1):
            full_name = str(cell_value(ws.cell(row, 2).value))
            if full_name and (full_name.lower() == lowered or full_name.split(".")[-1].lower() == lowered):
                return row
        return 0

    def _delete_definition_group(self, file_name: str, name: str) -> None:
        path = self.data_dir / file_name
        wb = load_workbook(path)
        lowered = name.lower()
        for ws in wb.worksheets:
            row = self._definition_start_row(ws, lowered)
            if row:
                count = 1
                while row + count <= ws.max_row and not cell_value(ws.cell(row + count, 2).value):
                    count += 1
                ws.delete_rows(row, count)
                wb.save(path)
                wb.close()
                return
        wb.close()
        die(f"definition not found: {name}")

    @staticmethod
    def _field_col(ws, field_name: str) -> int:
        for col in range(1, ws.max_column + 1):
            if str(cell_value(ws.cell(1, col).value)) == field_name:
                return col
        return 0


def print_json(value: Any) -> None:
    print(json.dumps(value, ensure_ascii=False, indent=2))


def main() -> None:
    parser = argparse.ArgumentParser(description="Tyou Luban Excel helper")
    parser.add_argument("--data-dir", default=str(DEFAULT_DATA_DIR))
    parser.add_argument("--write", action="store_true", help="allow write operations")
    sub = parser.add_subparsers(dest="command", required=True)

    table = sub.add_parser("table")
    table_sub = table.add_subparsers(dest="table_command", required=True)
    table_sub.add_parser("list")
    tget = table_sub.add_parser("get")
    tget.add_argument("name")
    tadd = table_sub.add_parser("add")
    tadd.add_argument("name")
    tadd.add_argument("--fields", required=True)
    tdel = table_sub.add_parser("delete")
    tdel.add_argument("name")

    field = sub.add_parser("field")
    field_sub = field.add_subparsers(dest="field_command", required=True)
    flist = field_sub.add_parser("list")
    flist.add_argument("table")
    fadd = field_sub.add_parser("add")
    fadd.add_argument("table")
    fadd.add_argument("name")
    fadd.add_argument("--type", required=True)
    fadd.add_argument("--comment", default="")
    fupdate = field_sub.add_parser("update")
    fupdate.add_argument("table")
    fupdate.add_argument("name")
    fupdate.add_argument("--new-name", default="")
    fupdate.add_argument("--type", default="")
    fupdate.add_argument("--comment", default="")
    fdel = field_sub.add_parser("delete")
    fdel.add_argument("table")
    fdel.add_argument("name")

    row = sub.add_parser("row")
    row_sub = row.add_subparsers(dest="row_command", required=True)
    rlist = row_sub.add_parser("list")
    rlist.add_argument("table")
    rlist.add_argument("--start", type=int, default=0)
    rlist.add_argument("--limit", type=int, default=20)
    rquery = row_sub.add_parser("query")
    rquery.add_argument("table")
    rquery.add_argument("--conditions", required=True)
    rquery.add_argument("--limit", type=int, default=20)
    radd = row_sub.add_parser("add")
    radd.add_argument("table")
    radd.add_argument("--data", required=True)
    rupdate = row_sub.add_parser("update")
    rupdate.add_argument("table")
    rupdate.add_argument("row", type=int)
    rupdate.add_argument("--data", required=True)
    rdel = row_sub.add_parser("delete")
    rdel.add_argument("table")
    rdel.add_argument("row", type=int)

    enum = sub.add_parser("enum")
    enum_sub = enum.add_subparsers(dest="enum_command", required=True)
    enum_sub.add_parser("list")
    eget = enum_sub.add_parser("get")
    eget.add_argument("name")
    eadd = enum_sub.add_parser("add")
    eadd.add_argument("name")
    eadd.add_argument("--items", required=True)
    eadd.add_argument("--comment", default="")
    eadd.add_argument("--flags", action="store_true")
    eupdate = enum_sub.add_parser("update")
    eupdate.add_argument("name")
    eupdate.add_argument("--comment", default="")
    eupdate.add_argument("--flags", choices=["true", "false"], default=None)
    edel = enum_sub.add_parser("delete")
    edel.add_argument("name")

    bean = sub.add_parser("bean")
    bean_sub = bean.add_subparsers(dest="bean_command", required=True)
    bean_sub.add_parser("list")
    bget = bean_sub.add_parser("get")
    bget.add_argument("name")
    badd = bean_sub.add_parser("add")
    badd.add_argument("name")
    badd.add_argument("--fields", default="")
    badd.add_argument("--parent", default="")
    badd.add_argument("--comment", default="")
    bupdate = bean_sub.add_parser("update")
    bupdate.add_argument("name")
    bupdate.add_argument("--parent", default="")
    bupdate.add_argument("--comment", default="")
    bdel = bean_sub.add_parser("delete")
    bdel.add_argument("name")

    validate = sub.add_parser("validate")
    validate.add_argument("table", nargs="?")
    validate.add_argument("--all", action="store_true")

    ref = sub.add_parser("ref")
    ref.add_argument("name")

    args = parser.parse_args()
    helper = LubanHelper(data_dir(args))

    if args.command == "table":
        if args.table_command == "list":
            print_json(helper.list_tables())
        elif args.table_command == "get":
            print_json(helper.table_info(args.name))
        elif args.table_command == "add":
            require_write(args)
            helper.add_table(args.name, parse_fields(args.fields))
            print_json({"ok": True})
        elif args.table_command == "delete":
            require_write(args)
            helper.delete_table(args.name)
            print_json({"ok": True})
    elif args.command == "field":
        if args.field_command == "list":
            print_json(helper.fields(args.table))
        elif args.field_command == "add":
            require_write(args)
            helper.add_field(args.table, args.name, args.type, args.comment)
            print_json({"ok": True})
        elif args.field_command == "update":
            require_write(args)
            helper.update_field(args.table, args.name, args.new_name, args.type, args.comment)
            print_json({"ok": True})
        elif args.field_command == "delete":
            require_write(args)
            helper.delete_field(args.table, args.name)
            print_json({"ok": True})
    elif args.command == "row":
        if args.row_command == "list":
            print_json(helper.rows(args.table, args.start, args.limit))
        elif args.row_command == "query":
            print_json(helper.query_rows(args.table, load_json_arg(args.conditions), args.limit))
        elif args.row_command == "add":
            require_write(args)
            helper.add_row(args.table, load_json_arg(args.data))
            print_json({"ok": True})
        elif args.row_command == "update":
            require_write(args)
            helper.update_row(args.table, args.row, load_json_arg(args.data))
            print_json({"ok": True})
        elif args.row_command == "delete":
            require_write(args)
            helper.delete_row(args.table, args.row)
            print_json({"ok": True})
    elif args.command == "enum":
        if args.enum_command == "list":
            print_json(helper.list_enums())
        elif args.enum_command == "get":
            print_json(helper.get_enum(args.name))
        elif args.enum_command == "add":
            require_write(args)
            helper.add_enum(args.name, parse_enum_items(args.items), args.comment, args.flags)
            print_json({"ok": True})
        elif args.enum_command == "update":
            require_write(args)
            flags = None if args.flags is None else args.flags == "true"
            helper.update_enum(args.name, args.comment, flags)
            print_json({"ok": True})
        elif args.enum_command == "delete":
            require_write(args)
            helper.delete_enum(args.name)
            print_json({"ok": True})
    elif args.command == "bean":
        if args.bean_command == "list":
            print_json(helper.list_beans())
        elif args.bean_command == "get":
            print_json(helper.get_bean(args.name))
        elif args.bean_command == "add":
            require_write(args)
            helper.add_bean(args.name, parse_fields(args.fields), args.parent, args.comment)
            print_json({"ok": True})
        elif args.bean_command == "update":
            require_write(args)
            helper.update_bean(args.name, args.parent, args.comment)
            print_json({"ok": True})
        elif args.bean_command == "delete":
            require_write(args)
            helper.delete_bean(args.name)
            print_json({"ok": True})
    elif args.command == "validate":
        if args.all or not args.table:
            print_json(helper.validate_all())
        else:
            print_json(helper.validate_table(args.table))
    elif args.command == "ref":
        print_json(helper.ref(args.name))


if __name__ == "__main__":
    main()
